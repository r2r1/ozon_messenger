import logging
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path

logger = logging.getLogger(__name__)

class ExcelExporter:
    def __init__(self, output_dir: Path, filename: str):
        self.output_dir = output_dir
        self.filename = filename
        self.filepath = output_dir / f"{filename}.xlsx"
    
    def export_results(self, data: dict, selected_fields: list = None) -> bool:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active

            # Поддерживаем 2 формата данных:
            # 1) products: [{..., seller: {...}}]
            # 2) sellers: [{...}]
            is_sellers = 'sellers' in data
            rows = data.get('sellers', []) if is_sellers else data.get('products', [])

            if is_sellers:
                ws.title = "Ozon Sellers"
                field_mapping = {
                    'seller_id': ('ID продавца', lambda s: s.get('seller_id', '')),
                    'seller_name': ('Продавец', lambda s: s.get('seller_name', '')),
                    'company_name': ('Название компании', lambda s: s.get('company_name', '')),
                    'inn': ('ИНН', lambda s: s.get('inn', '')),
                    'orders_count': ('Заказов', lambda s: s.get('orders_count', '')),
                    'reviews_count': ('Отзывов', lambda s: s.get('reviews_count', '')),
                    'average_rating': ('Рейтинг', lambda s: s.get('average_rating', '')),
                    'working_time': ('Работает с', lambda s: s.get('working_time', '')),
                    'seller_link': ('Ссылка продавца', lambda s: s.get('seller_link', '')),
                }
                default_fields = ['seller_name', 'company_name', 'seller_link', 'orders_count', 'average_rating', 'reviews_count', 'inn', 'working_time']
            else:
                ws.title = "Ozon Products"
                field_mapping = {
                    'article': ('Артикул', lambda p: p.get('article', '')),
                    'name': ('Название товара', lambda p: p.get('name', '')),
                    'seller_name': ('Продавец', lambda p: p.get('seller', {}).get('name', '')),
                    'company_name': ('Название компании', lambda p: p.get('seller', {}).get('company_name', '')),
                    'inn': ('ИНН', lambda p: p.get('seller', {}).get('inn', '')),
                    'card_price': ('Цена карты', lambda p: p.get('card_price', 0)),
                    'price': ('Цена', lambda p: p.get('price', 0)),
                    'original_price': ('Старая цена', lambda p: p.get('original_price', 0)),
                    'product_url': ('Ссылка товара', lambda p: p.get('product_url', '')),
                    'image_url': ('Изображение', lambda p: p.get('image_url', '')),
                    'orders_count': ('Заказов', lambda p: p.get('seller', {}).get('orders_count', '')),
                    'reviews_count': ('Отзывов', lambda p: p.get('seller', {}).get('reviews_count', '')),
                    'average_rating': ('Рейтинг', lambda p: p.get('seller', {}).get('average_rating', '')),
                    'working_time': ('Работает с', lambda p: p.get('seller', {}).get('working_time', ''))
                }
                default_fields = ['name', 'company_name', 'product_url', 'image_url']
            
            # Используем выбранные поля или все по умолчанию
            if selected_fields:
                included_fields = [field for field in selected_fields if field in field_mapping]
                if not included_fields:
                    included_fields = list(default_fields)
                headers = [field_mapping[field][0] for field in included_fields]
                field_extractors = [field_mapping[field][1] for field in included_fields]
            else:
                included_fields = list(default_fields)
                headers = [field_mapping[field][0] for field in default_fields]
                field_extractors = [field_mapping[field][1] for field in default_fields]
            
            # Стили
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            data_font = Font(name='Arial', size=10)
            data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Данные
            hyperlink_cols = set()
            if 'seller_link' in included_fields:
                hyperlink_cols.add(included_fields.index('seller_link') + 1)
            if (not is_sellers) and 'product_url' in included_fields:
                hyperlink_cols.add(included_fields.index('product_url') + 1)

            for row_idx, item in enumerate(rows, 2):
                row_data = [extractor(item) for extractor in field_extractors]

                # Цветовая индикация успешности (только для products, если есть флаг success)
                success_flag = bool(item.get('success')) if isinstance(item, dict) else True
                fill = (
                    PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    if success_flag
                    else PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                )

                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = border

                    if col_idx in hyperlink_cols and isinstance(value, str) and value.startswith(("http://", "https://")):
                        cell.hyperlink = value
                        cell.style = "Hyperlink"

                    # Красим только неуспешные items (актуально для products)
                    if isinstance(item, dict) and item.get('success') is False:
                        cell.fill = fill
            
            # Ширина колонок (адаптивная)
            default_widths = {
                'Артикул': 12,
                'Название товара': 40,
                'Продавец': 25,
                'Название компании': 30,
                'ИНН': 15,
                'Цена карты': 12,
                'Цена': 12,
                'Старая цена': 12,
                'Ссылка товара': 50,
                'Ссылка продавца': 50,
                'Изображение': 50,
                'ID продавца': 14,
                'Заказов': 12,
                'Отзывов': 12,
                'Рейтинг': 12,
                'Работает с': 15,
            }
            
            for col, header in enumerate(headers, 1):
                width = default_widths.get(header, 15)
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # Высота строк
            for r in range(1, len(rows) + 2):
                ws.row_dimensions[r].height = 20
            
            # Фильтр и заморозка
            if rows:
                ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
            ws.freeze_panes = "A2"
            
            wb.save(self.filepath)
            logger.info(f"Excel файл сохранен: {self.filepath}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка экспорта в Excel: {e}")
            return False